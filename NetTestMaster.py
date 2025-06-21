# ===================== 依赖检查区 =====================
# 本区用于检查脚本所需的所有第三方库是否已安装。
# 如果缺失依赖，则提示用户是否安装。

def check_dependencies():
    # 依赖检测用的 import 名和 pip 安装名的映射
    pip_names = {
        'ping3': 'ping3',
        'pandas': 'pandas',
        'wcwidth': 'wcwidth',
        'dns': 'dnspython',      # 检测 import dns，安装 dnspython
        'openpyxl': 'openpyxl'
    }
    missing_dependencies = []
    for import_name, pip_name in pip_names.items():
        try:
            __import__(import_name)
        except ImportError:
            missing_dependencies.append((import_name, pip_name))

    if missing_dependencies:
        print("检测到以下缺失依赖:")
        for import_name, pip_name in missing_dependencies:
            print(f"  - {pip_name}")
        choice = input("是否自动安装缺失依赖？(yes/no): ").strip().lower()
        if choice == 'yes':
            import subprocess
            for import_name, pip_name in missing_dependencies:
                subprocess.check_call(["pip", "install", pip_name])
            print("\n缺失依赖已安装完成，请重新运行脚本。")
            exit(0)
        else:
            print("请安装缺失依赖后再运行脚本。")
            exit(1)

# 在脚本开始时检查依赖
check_dependencies()

# ===================== 导入依赖区 =====================
# 本区用于导入脚本所需的所有第三方库和标准库。
# 包括异步IO、线程池、数据处理、时间、操作系统、网络、Excel美化等。

# 标准库
import asyncio           # 异步IO，用于并发调度主流程
import time              # 时间处理，获取时间戳等
import os                # 操作系统相关，如路径、文件操作
import json              # JSON序列化与反序列化，断点续传用
import socket            # 网络通信，UDP/TCP延迟测试和本地IP获取
import threading         # 线程锁，用于保护全局计数器total_sent的并发安全
import ipaddress         # 用于IP地址合法性判断

# 第三方库
import pandas as pd           # 数据处理与Excel导出
from ping3 import ping        # ICMP Ping库，用于ICMP协议延迟测试
from wcwidth import wcswidth  # 字符宽度计算，表格美观对齐
from concurrent.futures import ThreadPoolExecutor  # 线程池，用于多线程并发

script_start_time = time.time()  # 记录脚本启动时刻

# ===================== 工具函数区 =====================
# 本区包含通用的工具函数，如获取桌面路径、时间戳、IP判断等。

def get_desktop():
    """获取当前用户桌面路径"""
    return os.path.join(os.path.expanduser('~'), 'Desktop')

def get_timestamp():
    """获取当前时间戳字符串，格式为yyyyMMdd_HHmmss"""
    return time.strftime('%Y%m%d_%H%M%S')

def get_export_path():
    """自动生成带时间戳的Excel导出路径"""
    return os.path.join(get_desktop(), f'NetTest_{get_timestamp()}.xlsx')

def is_ip(address):
    """
    判断一个字符串是否为合法IP地址。
    返回True表示是合法IP，False表示不是。
    """
    try:
        ipaddress.ip_address(address)
        return True
    except ValueError:
        return False

def get_protocol_default_packet_size(protocol):
    """
    获取指定协议的推荐默认包大小（字节）。
    支持ICMP、UDP、TCP、DNS等协议。
    """
    proto = protocol.upper() if protocol else "UDP"
    if proto.startswith("ICMP"):
        return 56  # ICMP协议推荐包大小
    elif proto.startswith("UDP"):
        return 64  # UDP协议推荐包大小
    elif proto.startswith("TCP"):
        return 64  # TCP协议推荐包大小
    elif proto == "DNS":
        return 512 # DNS协议推荐包大小
    else:
        return 64  # 其他协议默认64字节

def get_packet_size(protocol):
    """
    根据配置返回当前测试所用包大小：
    - int类型：固定包大小
    - "default"：协议推荐默认包大小
    - "auto"：动态调整（当前等价于default，后续如需根据网络状况自适应可在此扩展）
    """
    pkt_cfg = config.get("packet_size", "default")
    if isinstance(pkt_cfg, int):
        return pkt_cfg
    if pkt_cfg == "default":
        return get_protocol_default_packet_size(protocol)
    if pkt_cfg == "auto":
        # 当前'auto'等价于'default'，如需自适应可扩展此分支
        return get_protocol_default_packet_size(protocol)
    return 64  # 兜底默认值

# ===================== 配置参数区 =====================
# 本区用于集中管理所有脚本参数，包括基础参数、测试参数、导出参数、显示参数、断点续传参数等。
config = {
    # ===== 基础参数 =====
    "total_scan_time": 60,             # 总扫描时间限制（秒），超时强制终止
    "concurrent_threads": 100,         # 并发线程数
    "timeout": 0.5,                    # 单次请求超时时间（秒）
    "protocol_type": "TCP",            # 测试协议类型：ICMP、UDP、TCP
    "enable_ipv6": False,              # 是否启用IPv6测试
    "use_local_ip": True,              # 是否使用本地IP测试
    "resolved_ip_type": "A",           # 解析IP类型：'A'（IPv4）、'AAAA'（IPv6）、'auto'（随enable_ipv6）

    # ===== 测试目标参数 =====
    "test_count_per_dns": 3,           # 每个目标测试次数
    "test_domain": "www.baidu.com",    # 指定测试域名（留空则按协议类型测试IP）
    "total_test_count": None,          # 限制总测试次数（None不限制）

    # ===== 导出参数 =====yes
    "export_to_desktop": True,         # 是否导出到桌面
    "export_to_script_dir": False,     # 是否导出到脚本同路径下
    "export_path": get_export_path(),  # 兜底导出路径（自动生成）

    # ===== 显示与导出字段参数 =====
    "top_n": 30,                       # 显示/导出前N个结果
    "show_config": True,               # 是否输出配置区内容
    "show_recv_sent": True,            # 是否输出Recv/Sent数据
    "show_loss_rate": True,            # 是否输出Loss Rate(%)数据
    "show_protocol": True,             # 是否输出Protocol数据
    "show_country": True,              # 是否输出Country数据
    "show_packet_size": True,          # 是否输出测试包大小数据

    # ===== 断点续传参数 =====
    "enable_resume": False,            # 是否启用断点续传功能
    "resume_file": "scan_resume.json", # 断点续传记录文件

    # ===== 协议测试数据包大小参数 =====
    "packet_size": "auto",       # 测试包大小（字节）：int为自定义固定，"default"为协议推荐，"auto"为动态调整（当为int时，请注意数值不要加引号）
    "show_packet_size": True,    # 是否输出测试包大小数据
}

# ========== 导出路径自动适配逻辑 ==========
# 优先级：桌面 > 脚本同路径 > 其它
if config.get("export_to_desktop", False):
    config["export_path"] = os.path.join(get_desktop(), f'NetTest_{get_timestamp()}.xlsx')
elif config.get("export_to_script_dir", False):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config["export_path"] = os.path.join(script_dir, f'NetTest_{get_timestamp()}.xlsx')
# 否则保持原有export_path

# ===================== Excel导出格式配置区 =====================
# 本区用于统一配置导出Excel文件的表头、数据内容、各列宽度和对齐方式等格式，方便后续美化和维护。
excel_column_config = {
    'Index':        {'width': 10, 'align': 'center'},   # 序号列，适合居中
    'Name':         {'width': 20, 'align': 'center'},   # DNS名称列，适合居中
    'Address':      {'width': 20, 'align': 'center'},   # IP地址列，适合居中
    'Latency(ms)':  {'width': 15, 'align': 'center'},   # 延迟列，适合居中
    'Resolved IP':  {'width': 20, 'align': 'center'},   # 解析结果IP列，适合居中
    'Packet Size':  {'width': 13, 'align': 'center'},   # 测试包大小，左对齐
    'Recv/Sent':    {'width': 15, 'align': 'center'},   # 收发包数列，适合居中
    'Loss Rate(%)': {'width': 15, 'align': 'center'},   # 丢包率列，适合居中
    'Protocol':     {'width': 15, 'align': 'center'},   # 协议类型列，适合居中
    'Country':      {'width': 15, 'align': 'center'},   # 国家列，适合居中
}
# width: 列宽，单位为Excel的字符宽度（1为一个标准字符宽度）
# align: 列内容对齐方式，可选值：
#   'center' —— 居中对齐，适合序号、协议、数值等
#   'left'   —— 左对齐，适合文本型内容如名称、国家
#   'right'  —— 右对齐适合数值型内容如延迟、丢包率、IP等
# 实际应用时根据美观和数据类型选择合适的对齐方式。

excel_export_format = {
    'header': {
        'row_height': 30,                # 表头行高（像素）
        'font_name': 'OPPO Sans 4.0',    # 表头字体
        'font_size': 12,                 # 表头字号
        'align': 'center',               # 表头居中对齐
    },
    'data': {
        'font_name': 'OPPO Sans 4.0 light', # 数据内容字体
        'font_size': 10,                    # 数据内容字号
        'row_height': 20,                   # 数据内容行高
    }
}

# ===================== 预定义的公共DNS列表 =====================
# 本区定义常用的公共DNS服务器IP，便于测速和展示，可根据需要增删
DNS_NAMES = {
    '8.8.8.8':           'Google DNS',
    '8.8.4.4':           'Google DNS',
    '1.1.1.1':           'Cloudflare',
    '223.5.5.5':         'AliDNS',
    '223.6.6.6':         'AliDNS',
    '119.29.29.29':      '腾讯云DNS',
    '1.0.0.1':           'Cloudflare',
    '9.9.9.9':           'Quad9',
    '149.112.112.112':   'Quad9',
    '208.67.222.222':    'OpenDNS',
    '208.67.220.220':    'OpenDNS',
    '94.140.14.14':      'AdGuard',
    '94.140.15.15':      'AdGuard',
    '185.228.168.9':     'CleanBrowsing',
    '185.228.169.9':     'CleanBrowsing',
    '64.6.64.6':         'Verisign',
    '64.6.65.6':         'Verisign',
    '84.200.69.80':      'DNS.WATCH',
    '84.200.70.40':      'DNS.WATCH',
    '8.20.247.20':       'Comodo',
    '199.85.126.10':     'Norton',
    '156.154.70.1':      'SafeDNS',
    '45.90.28.0':        'NextDNS',
    '76.76.19.19':       'Alternate DNS',
    '76.223.122.150':    'Alternate DNS',
}

# ===================== DNS国家映射表 =====================
# 用于显示每个DNS服务器的归属国家
DNS_COUNTRY = {
    '8.8.8.8':           '美国',
    '8.8.4.4':           '美国',
    '1.1.1.1':           '澳大利亚',
    '1.0.0.1':           '澳大利亚',
    '223.5.5.5':         '中国',
    '223.6.6.6':         '中国',
    '119.29.29.29':      '中国',
    '9.9.9.9':           '美国',
    '149.112.112.112':   '美国',
    '208.67.222.222':    '美国',
    '208.67.220.220':    '美国',
    '94.140.14.14':      '德国',
    '94.140.15.15':      '德国',
    '185.228.168.9':     '以色列',
    '185.228.169.9':     '以色列',
    '64.6.64.6':         '美国',
    '64.6.65.6':         '美国',
    '84.200.69.80':      '德国',
    '84.200.70.40':      '德国',
    '8.20.247.20':       '美国',
    '199.85.126.10':     '美国',
    '156.154.70.1':      '美国',
    '45.90.28.0':        '瑞士',
    '76.76.19.19':       '美国',
    '76.223.122.150':    '美国',
}

# ===================== 断点续传功能区 =====================
# 本区实现断点续传相关的加载、保存、清理功能，便于中断后恢复测速进度。
def load_resume():
    """
    加载断点续传记录，返回已完成的结果和索引集合。
    参数变动时自动清理断点文件。
    """
    resume_path = config.get("resume_file", "scan_resume.json")
    key_params = {
        'test_count_per_dns': config.get('test_count_per_dns'),
        'test_domain': config.get('test_domain'),
        'protocol_type': config.get('protocol_type'),
        'enable_ipv6': config.get('enable_ipv6'),
    }
    if config.get("enable_resume") and os.path.exists(resume_path):
        try:
            with open(resume_path, "r") as f:
                data = json.load(f)
                old_params = data.get('params', {})
                # 检查关键参数是否有变动
                if any(key_params.get(k) != old_params.get(k) for k in key_params):
                    os.remove(resume_path)
                    return { }, set()
                return data.get('results', {}), set(data.get('finished_idx', []))
        except Exception:
            # 文件损坏等异常也清理
            os.remove(resume_path)
            return {}, set()
    return {}, set()

def save_resume(results, finished_idx):
    """
    保存断点续传记录，支持断点恢复。
    记录关键参数快照。
    """
    if config.get("enable_resume"):
        key_params = {
            'test_count_per_dns': config.get('test_count_per_dns'),
            'test_domain': config.get('test_domain'),
            'protocol_type': config.get('protocol_type'),
            'enable_ipv6': config.get('enable_ipv6'),
        }
        with open(config["resume_file"], "w") as f:
            json.dump({'results': results, 'finished_idx': list(finished_idx), 'params': key_params}, f)

def clear_resume():
    """
    测速前自动清理无用断点文件，防止脏数据影响新测试。
    仅在启用断点续传功能时生效。
    """
    if config.get("enable_resume") and os.path.exists(config["resume_file"]):
        os.remove(config["resume_file"])

# ===================== 域名解析相关区 =====================
# 本区实现域名解析相关函数，包括本地和指定DNS服务器解析。
def resolve_domain(domain, ipv6=False):
    """
    使用dnspython库解析域名，返回第一个A/AAAA记录IP。
    resolved_ip_type控制解析类型：A/AAAA/auto
    """
    import dns.resolver
    resolver = dns.resolver.Resolver()
    resolved_type = config.get('resolved_ip_type', 'auto')
    if resolved_type == 'A':
        qtype = 'A'
    elif resolved_type == 'AAAA':
        qtype = 'AAAA'
    else:
        qtype = 'AAAA' if ipv6 else 'A'
    answer = resolver.resolve(domain, qtype)
    for rdata in answer:
        return rdata.address

def resolve_domain_with_dns(domain, dns_server, ipv6=False):
    """使用指定DNS服务器解析域名，返回第一个A/AAAA记录IP。异常时返回None。"""
    try:
        import dns.resolver
        resolver = dns.resolver.Resolver()
        resolver.nameservers = [dns_server]
        resolver.timeout = config.get('timeout', 2)
        resolver.lifetime = config.get('timeout', 2)
        resolved_type = config.get('resolved_ip_type', 'auto')
        if resolved_type == 'A':
            qtype = 'A'
        elif resolved_type == 'AAAA':
            qtype = 'AAAA'
        else:
            qtype = 'AAAA' if ipv6 else 'A'
        answer = resolver.resolve(domain, qtype)
        for rdata in answer:
            return rdata.address
    except Exception:
        return None

# ===================== 核心测速与并发调度区 =====================
# 本区实现核心测速逻辑，包括ICMP/UDP/TCP/DNS测速、并发调度、统计等。
# 包括test_ping、run_concurrent_ping等函数。
# 检查 protocol_type 合法性
VALID_PROTOCOLS = {"ICMP", "UDP", "TCP"}
if config.get("protocol_type", "ICMP") not in VALID_PROTOCOLS:
    print(f"[警告] 配置的 protocol_type '{config.get('protocol_type')}' 非法，已自动回退为 ICMP。"); config["protocol_type"] = "ICMP"

# addresses 生成逻辑，根据是否指定test_domain决定目标列表
addresses = DNS_NAMES  # 目前无论 test_domain 是否为空都用同一组目标，可扩展为自定义目标

# ===================== 并发执行Ping测试 =====================
resume_lock = threading.Lock()  # 断点续传写入锁
async def run_concurrent_ping():
    """
    并发执行所有DNS的Ping测试，支持断点续传。
    每完成一个目标就保存一次进度。
    返回每个DNS的测速结果。
    """
    results, finished_idx = load_resume()
    results = dict(results)  # 确保可变
    finished_idx = set(finished_idx)
    with ThreadPoolExecutor(max_workers=config["concurrent_threads"]) as executor:
        loop = asyncio.get_event_loop()
        futures = {}
        for address in addresses.keys():
            if address in finished_idx:
                continue
            futures[address] = loop.run_in_executor(executor, test_ping, address)
        for address, future in futures.items():
            result = await future
            with resume_lock:
                results[address] = result
                finished_idx.add(address)
                save_resume(results, finished_idx)
    # 返回所有目标的测速结果（包括已完成和新测的）
    all_results = [results[addr] for addr in addresses.keys() if addr in results]
    return all_results

# ===================== 表格格式配置区 =====================
# 本区用于配置终端表格输出的各列间隙和对齐方式，便于美观和自定义。
# 各列间隙（单位：空格数）
index_left_margin      = 1   # Index列左边距
index_name_gap         = 3   # Index与Name之间的间隙
name_address_gap       = 3   # Name与Address之间的间隙
address_latency_gap    = 3   # Address与Latency(ms)之间的间隙
latency_resolvedip_gap = 3   # Latency(ms)与Resolved IP之间的间隙
packet_size_gap       = 3    # Resolved IP与Packet Size之间的间隙
packet_recvsent_gap   = 4    # Packet Size与Recv/Sent之间的间隙
recvsent_lossrate_gap  = 3   # Recv/Sent与Loss Rate(%)之间的间隙
lossrate_protocol_gap  = 3   # Loss Rate(%)与Protocol之间的间隙
protocol_country_gap   = 3   # Protocol与Country之间的间隙
country_right_margin   = 1   # Country列右边距

# 各列对齐方式
index_align      = 'center'  # Index列，居中对齐
name_align       = 'left'    # Name列，左对齐
address_align    = 'left'    # Address列，左对齐
latency_align    = 'left'    # Latency(ms)列，左对齐
resolvedip_align = 'right'   # Resolved IP列，右对齐
packet_size_align = 'right'  # Packet Size列，右对齐
recvsent_align   = 'center'  # Recv/Sent列，居中对齐
lossrate_align   = 'right'   # Loss Rate(%)列，右对齐
protocol_align   = 'center'  # Protocol列，居中对齐
country_align    = 'center'  # Country列，居中对齐

# ===================== 主函数 =====================
async def main():
    start_time = script_start_time  # 记录脚本启动到测试完成的总用时
    """
    主流程：
    1. 输出配置区内容（可选）
    2. 输出测试开始提示
    3. 并发执行测速
    4. 排序、输出表格
    5. 导出Excel（可选）
    test_domain 配置优先级高于 protocol_type，指定 test_domain 时强制使用 DNS 查询协议。
    """
    # 测试前自动清理无用断点文件
    local_ip = get_local_ip() if config.get('use_local_ip', True) else None
    total_dns  = len(addresses)
    # 输出配置区内容（如开关为True）
    if config.get('show_config', True):
        print("当前配置:")
        for k, v in config.items():
            print(f"  {k}: {v}")
    # 输出测试开始提示
    if config.get('test_domain'):
        if local_ip:
            print(f"共发现 {total_dns} 个DNS服务器，使用本地IP:{local_ip} 指定地址:{config['test_domain']}")
        else:
            print(f"共发现 {total_dns} 个DNS服务器，指定地址:{config['test_domain']}")
        print("开始测试，请稍等......")
    else:
        if local_ip:
            print(f"共发现 {total_dns} 个DNS服务器，使用本地IP: {local_ip}")
        else:
            print(f"共发现 {total_dns} 个DNS服务器")
        print("开始测试，请稍等......")
    # 执行测速
    results = await run_concurrent_ping()
    # 按延迟从低到高排序，超时排最后
    results = sorted(results, key=lambda x: (float('inf') if isinstance(x[1], str) else x[1]))
    # 无论超时与否，始终输出和导出前top_n个
    results_to_show = results[:config["top_n"]]
    # 读取配置项（提前到所有用到 show_recv_sent 之前）
    show_protocol = config.get("show_protocol")
    if show_protocol is None:
        show_protocol = True
    show_loss_rate = config.get("show_loss_rate")
    if show_loss_rate is None:
        show_loss_rate = True
    show_country = config.get("show_country")
    if show_country is None:
        show_country = True
    show_recv_sent = config.get("show_recv_sent")
    if show_recv_sent is None:
        show_recv_sent = True
    show_packet_size = config.get("show_packet_size")

    # 表格输出格式设置
    max_idx         = len(results_to_show)
    idx_width       = max(len(str(max_idx)), len('Index')) + 2  # Index列宽
    name_width      = max(14, max([wcswidth(DNS_NAMES.get(addr, '-') or '-') for addr, *_ in results_to_show] + [wcswidth('Name')]) + 2)
    # 修正Protocol列宽度为表头和所有内容的最大宽度
    protocol_width  = max(8, max([wcswidth(str(row[2])) for row in results_to_show] + [wcswidth('Protocol')]))
    country_width   = max(8, max([wcswidth(DNS_COUNTRY.get(addr, '-') or '-') for addr, *_ in results_to_show] + [wcswidth('Country')]) + 2)
    loss_rate_strs = [str(row[3]) + '%' for row in results_to_show]
    max_loss_rate_width = max([wcswidth(s) for s in loss_rate_strs] + [wcswidth('Loss Rate(%)')])
    loss_width = max(14, max_loss_rate_width + 2)  # 14为最小宽度
    # Recv/Sent 列宽度自适应（内容和表头最大宽度）
    if show_recv_sent:
        recvsent_col_width = max([wcswidth(f"{row[5]}/{row[4]}") for row in results_to_show] + [wcswidth('Recv/Sent')])
    else:
        recvsent_col_width = 12

    # Packet Size 列宽度基于表头宽度和所有数据的最大宽度
    if show_packet_size:
        packet_size_col_width = max([wcswidth(str(row[7])) for row in results_to_show] + [wcswidth('Packet Size')])
    else:
        packet_size_col_width = 13

    # Resolved IP 最大显示宽度与 packet_size_gap 动态关联
    base_resolved_ip_width = 15
    resolved_ip_max_width = base_resolved_ip_width + packet_size_gap

    # 动态拼接表头
    header = (
        ' ' * index_left_margin +
        f"{'Index':^{idx_width}}" +
        ' ' * index_name_gap +
        pad_right('Name', name_width) +
        ' ' * name_address_gap +
        f"{'Address':<20}" +
        ' ' * address_latency_gap +
        f"{'Latency(ms)':<14}"
    )
    if config.get("test_domain"):
        header += ' ' * latency_resolvedip_gap + f"{'Resolved IP':<{resolved_ip_max_width}}"
    if show_packet_size:
        if packet_size_align == 'right':
            header += ' ' * packet_size_gap + f"{'Packet Size':>{packet_size_col_width}}"
        elif packet_size_align == 'center':
            header += ' ' * packet_size_gap + pad_center('Packet Size', packet_size_col_width)
        else:
            header += ' ' * packet_size_gap + f"{'Packet Size':<{packet_size_col_width}}"
    if show_recv_sent:
        header += ' ' * packet_recvsent_gap + f"{'Recv/Sent':>{recvsent_col_width}}"
    if show_loss_rate:
        header += ' ' * recvsent_lossrate_gap + pad_center('Loss Rate(%)', loss_width)
    if show_protocol:
        header += ' ' * lossrate_protocol_gap + pad_center('Protocol', protocol_width)
    if show_country:
        header += ' ' * protocol_country_gap + pad_center('Country', country_width)
    header += ' ' * country_right_margin
    print("-" * wcswidth(header))
    print(header)
    print("-" * wcswidth(header))
    # 输出每一行测速结果
    # Resolved IP 最大显示宽度
    # resolved_ip_max_width 已经动态设置

    def fold_resolved_ip(ip, max_width):
        """如果IP字符串超出最大宽度则折叠显示（前缀...后缀），始终左对齐并占满max_width"""
        if not ip:
            return '-'.ljust(max_width)
        ip = str(ip)
        if wcswidth(ip) <= max_width:
            return ip.ljust(max_width)
        ellipsis = '...'
        remain = max_width - wcswidth(ellipsis)
        left = remain // 2
        right = remain - left
        prefix = ip[:left]
        suffix = ip[-right:] if right > 0 else ''
        folded = f"{prefix}{ellipsis}{suffix}"
        # 保证宽度
        return folded.ljust(max_width)

    for idx, row_data in enumerate(results_to_show, 1):
        address, latency, protocol, loss_rate, sent_count, received_count, resolved_ip, packet_size = row_data
        name = DNS_NAMES.get(address, '-')
        country = DNS_COUNTRY.get(address, '-')
        latency_str = f"{latency}" if isinstance(latency, (int, float)) else "Timeout"
        latency_str = pad_right(latency_str, 14)
        # 保留两位小数的百分比
        loss_rate_str = f"{loss_rate:.2f}%"
        right_aligned = loss_rate_str.rjust(max_loss_rate_width)
        row = (
            ' ' * index_left_margin +
            f"{str(idx):^{idx_width}}" +
            ' ' * index_name_gap +
            pad_right(name, name_width) +
            ' ' * name_address_gap +
            f"{address:<20}" +
            ' ' * address_latency_gap +
            latency_str
        )
        if config.get("test_domain"):
            row += ' ' * latency_resolvedip_gap + fold_resolved_ip(resolved_ip, resolved_ip_max_width)
        if show_packet_size:
            if packet_size_align == 'right':
                row += ' ' * packet_size_gap + f"{str(packet_size):>{packet_size_col_width}}"
            elif packet_size_align == 'center':
                row += ' ' * packet_size_gap + pad_center(str(packet_size), packet_size_col_width)
            else:
                row += ' ' * packet_size_gap + f"{str(packet_size):<{packet_size_col_width}}"
        if show_recv_sent:
            row += ' ' * packet_recvsent_gap + f"{(str(received_count) + '/' + str(sent_count)):>{recvsent_col_width}}"
        if show_loss_rate:
            row += ' ' * recvsent_lossrate_gap + pad_center(right_aligned, loss_width)
        if show_protocol:
            row += ' ' * lossrate_protocol_gap + pad_center(protocol, protocol_width)
        if show_country:
            row += ' ' * protocol_country_gap + pad_center(country, country_width)
        row += ' ' * country_right_margin
        print(row)
    # 测试完成提示
    total_time = time.time() - start_time
    h = int(total_time // 3600)
    m = int((total_time % 3600) // 60)
    s = int(total_time % 60)
    cs = int((total_time - int(total_time)) * 100)  # 百分秒
    print("\n测试完成！")
    print(f"总共用时：{h:02d}:{m:02d}:{s:02d}.{cs:02d}")
    # 输出断点续传记录文件路径
    if config.get('enable_resume'):
        resume_file_path = os.path.abspath(config.get('resume_file', 'scan_resume.json'))
        resume_dir = os.path.dirname(resume_file_path)
        print(f"断点续传记录文件(scan_resume.json)已导出到: {resume_dir}")
    if config.get('export_to_desktop') or config.get('export_to_script_dir'):
        export_path = config.get('export_path', 'NetTest.xlsx')
        # 整理数据
        df = pd.DataFrame([
            {
                "Index": idx,
                "Name": DNS_NAMES.get(address, '-'),
                "Address": address,
                "Latency(ms)": latency if isinstance(latency, (int, float)) else None,
                **({"Resolved IP": resolved_ip} if config.get("test_domain") else {}),
                **({"Packet Size": packet_size} if show_packet_size else {}),
                **({"Recv/Sent": f"{received_count}/{sent_count}"} if config.get("show_recv_sent") else {}),
                **({"Loss Rate(%)": loss_rate} if config.get("show_loss_rate") else {}),
                **({"Protocol": protocol} if config.get("show_protocol") else {}),
                **({"Country": DNS_COUNTRY.get(address, '-')} if config.get("show_country") else {}),
            }
            for idx, (address, latency, protocol, loss_rate, sent_count, received_count, resolved_ip, packet_size) in enumerate(results_to_show, 1)
        ])
        # 调整列顺序以与终端输出一致
        desired_columns = [
            "Index",
            "Name",
            "Address",
            "Latency(ms)",
        ]
        if config.get("test_domain"):
            desired_columns.append("Resolved IP")
        if show_packet_size:
            desired_columns.append("Packet Size")
        if config.get("show_recv_sent"):
            desired_columns.append("Recv/Sent")
        if config.get("show_loss_rate"):
            desired_columns.append("Loss Rate(%)")
        if config.get("show_protocol"):
            desired_columns.append("Protocol")
        if config.get("show_country"):
            desired_columns.append("Country")
        # 只保留实际存在的列
        df = df[[col for col in desired_columns if col in df.columns]]

        # Loss Rate(%)列设置为小数（如0.6），以便Excel百分比格式
        if "Loss Rate(%)" in df.columns:
            df["Loss Rate(%)"] = df["Loss Rate(%)"].apply(
                lambda x: float(str(x).replace('%',''))/100 if pd.notnull(x) and str(x).strip() != '' else None
            )

        # ====== 最小侵入式多路径导出实现 ======
        export_to_desktop = config.get("export_to_desktop", False)
        export_to_script_dir = config.get("export_to_script_dir", False)
        desktop_path = os.path.join(get_desktop(), f'NetTest_{get_timestamp()}.xlsx')
        script_dir = os.path.dirname(os.path.abspath(__file__))
        script_path = os.path.join(script_dir, f'NetTest_{get_timestamp()}.xlsx')
        export_results = []
        if export_to_desktop and export_to_script_dir:
            paths = [(desktop_path, 'desktop'), (script_path, 'script')]
        elif export_to_desktop:
            paths = [(desktop_path, 'desktop')]
        elif export_to_script_dir:
            paths = [(script_path, 'script')]
        else:
            paths = [(export_path, 'other')]
        for path, tag in paths:
            df.to_excel(path, index=False)
            import openpyxl
            from openpyxl.styles import Alignment, Font, Border, Side
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            header_cfg = excel_export_format['header']
            no_border = Border(left=Side(border_style=None), right=Side(border_style=None), top=Side(border_style=None), bottom=Side(border_style=None))
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(name=header_cfg['font_name'], size=header_cfg['font_size'], bold=False)
                cell.alignment = Alignment(horizontal=header_cfg['align'], vertical='center')
                cell.border = no_border  # 显式无边框
                ws.row_dimensions[1].height = header_cfg['row_height']
                width = excel_column_config.get(col_name, {}).get('width', 15)
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
            data_cfg = excel_export_format['data']
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
                ws.row_dimensions[row_idx].height = data_cfg.get('row_height', 20)
                for cell, col_name in zip(row, df.columns):
                    cell.font = Font(name=data_cfg.get('font_name', 'OPPO Sans 4.0 light'), size=data_cfg.get('font_size', 10))
                    align = excel_column_config.get(col_name, {}).get('align', 'center')
                    cell.alignment = Alignment(horizontal=align, vertical='center')
            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name == "Loss Rate(%)":
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.number_format = '0.00%'
            wb.save(path)
            export_results.append((os.path.basename(path), os.path.dirname(os.path.abspath(path)), tag))
        # 输出顺序：桌面优先，其次脚本目录
        for tag in ['desktop', 'script', 'other']:
            for export_file, export_dir, t in export_results:
                if t == tag:
                    print(f"测试结果({export_file})已导出到\"{export_dir}\"")

def pad_center(text, width):
    """辅助函数：将文本居中填充到指定宽度，支持宽字符，仅用于终端输出"""
    text = str(text)
    pad = width - wcswidth(text)
    left = pad // 2
    right = pad - left
    return ' ' * left + text + ' ' * right

def pad_right(text, width):
    """辅助函数：将文本右侧填充到指定宽度，支持宽字符，仅用于终端输出"""
    text = str(text)
    pad = width - wcswidth(text)
    return text + ' ' * pad

# ===================== 获取本地IP =====================
def get_local_ip():
    """获取本地IP地址，优先获取外网出口IP"""
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
    except Exception:
        ip = '127.0.0.1'
    finally:
        s.close()
    return ip

# ===================== DNS测速核心函数 =====================
# 在主流程前定义全局计数器
total_sent = 0
sent_lock = threading.Lock()

def udp_ping(address, timeout=2, ipv6=False):
    """对目标DNS服务器的53端口进行UDP延迟测试，返回秒（float）或None。支持IPv4/IPv6。使用标准DNS查询包。自动兼容dnspython 1.x和2.x。"""
    try:
        import dns.message
        import dns.query
        import inspect
        q = dns.message.make_query('www.example.com', 'A')
        start = time.time()
        # 检查udp()是否支持af参数
        if 'af' in inspect.signature(dns.query.udp).parameters:
            import socket
            response = dns.query.udp(q, address, timeout=timeout, af=socket.AF_INET6 if ipv6 else socket.AF_INET)
        else:
            response = dns.query.udp(q, address, timeout=timeout)
        end = time.time()
        return end - start
    except Exception:
        return None

def tcp_ping(address, timeout=2, ipv6=False):
    """对目标DNS服务器的53端口进行TCP连接延迟测试，返回秒（float）或None。支持IPv4/IPv6。"""
    try:
        family = socket.AF_INET6 if ipv6 else socket.AF_INET
        sock = socket.socket(family, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        start = time.time()
        sock.connect((address, 53))
        end = time.time()
        return end - start
    except Exception:
        return None
    finally:
        sock.close()

def test_ping(address):
    """
    对单个DNS服务器进行多次延迟测试，返回统计结果。
    支持ICMP/UDP/TCP和IPv6。
    返回：(address, 平均延迟/超时, 协议, 丢包率, 发送数, 接收数, 解析IP, 包大小)
    """
    global total_sent
    latencies = []
    protocol = None
    timeout_count = 0
    sent_count = 0
    ipv6 = config.get("enable_ipv6", False)
    resolved_ip = None
    packet_size = None
    for _ in range(config["test_count_per_dns"]):
        with sent_lock:
            if config.get("total_test_count") is not None and total_sent >= config["total_test_count"]:
                break
            total_sent += 1
        sent_count += 1
        try:
            if config.get("test_domain"):
                start = time.time()
                ip = resolve_domain_with_dns(config["test_domain"], address, ipv6)
                latency = (time.time() - start) if ip else None
                protocol = "DNS"
                if ip:
                    resolved_ip = ip
                packet_size = get_packet_size(protocol)
            elif config.get("protocol_type", "ICMP") == "ICMP":
                try:
                    if ipv6:
                        latency = None
                    else:
                        latency = ping(address, timeout=config["timeout"])
                except PermissionError:
                    print("[错误] ICMP 协议需要管理员权限，请以管理员身份运行脚本，或切换为 TCP/UDP 协议。"); return address, "权限不足", "ICMP", 100.0, sent_count, 0, resolved_ip, get_packet_size("ICMP")
                protocol = "ICMPv6" if ipv6 else "ICMP"
                packet_size = get_packet_size(protocol)
            elif config.get("protocol_type") == "UDP":
                latency = udp_ping(address, timeout=config["timeout"], ipv6=ipv6)
                protocol = "UDPv6" if ipv6 else "UDP"
                packet_size = get_packet_size(protocol)
            elif config.get("protocol_type") == "TCP":
                latency = tcp_ping(address, timeout=config["timeout"], ipv6=ipv6)
                protocol = "TCPv6" if ipv6 else "TCP"
                packet_size = get_packet_size(protocol)
            else:
                latency = None
                protocol = "Unknown"
                packet_size = get_packet_size(protocol)
            if latency is not None:
                latencies.append(latency * 1000)
            else:
                timeout_count += 1
        except Exception:
            timeout_count += 1
    total = len(latencies) + timeout_count if latencies or timeout_count else config["test_count_per_dns"]
    loss_rate = round(timeout_count / total * 100, 2) if total else 0
    received_count = len(latencies)
    if latencies:
        return address, round(sum(latencies) / len(latencies), 2), protocol, loss_rate, sent_count, received_count, resolved_ip, packet_size
    else:
        return address, "请求超时", protocol, loss_rate, sent_count, received_count, resolved_ip, packet_size

# ===================== 程序入口 =====================
# 仅当直接运行本文件时才会执行测速主流程
if __name__ == "__main__":
    try:
        asyncio.run(asyncio.wait_for(main(), timeout=config["total_scan_time"]))
    except asyncio.TimeoutError:
        print(f"\n[超时警告] 总扫描时间已达 {config['total_scan_time']} 秒，程序被强制终止。")